#!/usr/bin/env python3
"""
PDF -> Excel extractor for MOVE logistics-center brochures.

Each PDF page is a single (scanned / image-only) property sheet that follows a
fixed template:

    +-------------------------------------------------------------+
    | N. <창고명> (상온)                                           |
    | [building photo]   Location(map)        Site Plan           |
    | General Information            |  Space Availability table   |
    |   소재지 / 인근IC / 건폐율 ... |  (공실현황)                  |
    +-------------------------------------------------------------+

Because the pages carry no text layer, every value is recovered with OCR
(Tesseract, Korean + English). The General-Information table is located by
detecting its horizontal grid lines, then each value cell is OCR'd on its own
which is far more reliable than reading the whole block at once.

The script writes one Excel row per page with these columns:

    사용가능여부 · 중개인/임대인명 · 창고명 · 주소 · 행정구역_도 · 행정구역_시
    · 대지면적 · 연면적 · 건축면적 · 준공연도 · 공실현황(image) · 시설특이사항
    · 기타 · 정보확인일자 · 사진(image)

Usage:
    python main.py dsvtest2.pdf                       # -> dsvtest2.xlsx
    python main.py input.pdf -o out.xlsx --broker S1
    python main.py input.pdf --dpi 300
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import date

import fitz  # PyMuPDF
import numpy as np
import pytesseract
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image

# ---------------------------------------------------------------------------
# Layout constants (fractions of page width / height -> resolution independent)
# ---------------------------------------------------------------------------
OCR_LANG = "kor+eng"

# General-Information table search box (left half of the lower page).
GI_BOX = dict(x0=0.060, x1=0.302, y0=0.596, y1=0.880)
# x where the value column starts inside a GI row (label sits to the left).
GI_VALUE_X = 0.133
# Ordered field labels of the 9 GI rows.
GI_FIELDS = ["소재지", "인근IC", "건폐율/용적률", "대지면적",
             "연면적", "규모", "형태", "주차", "준공년도"]

# Title band (the "N. 창고명 (상온)" headline).
TITLE_BOX = dict(x0=0.056, x1=0.485, y0=0.142, y1=0.190)

# Building photo region (top-left), trimmed to its real content afterwards.
PHOTO_BOX = dict(x0=0.057, x1=0.306, y0=0.224, y1=0.572)

# Space-Availability table (공실현황) on the right.
SPACE_BOX = dict(x0=0.497, x1=0.978, y0=0.577, y1=0.893)

# Facility-spec / 비고 middle column (시설특이사항).
SPEC_BOX = dict(x0=0.304, x1=0.496, y0=0.603, y1=0.873)

# Canonical province name -> surface forms that may appear at the start of an
# address (longest first so the most specific prefix is stripped).
PROVINCE = [
    ("서울특별시", ["서울특별시", "서울"]),
    ("부산광역시", ["부산광역시", "부산"]),
    ("대구광역시", ["대구광역시", "대구"]),
    ("인천광역시", ["인천광역시", "인천"]),
    ("광주광역시", ["광주광역시", "광주"]),
    ("대전광역시", ["대전광역시", "대전"]),
    ("울산광역시", ["울산광역시", "울산"]),
    ("세종특별자치시", ["세종특별자치시", "세종시", "세종"]),
    ("경기도", ["경기도", "경기"]),
    ("강원특별자치도", ["강원특별자치도", "강원도", "강원"]),
    ("충청북도", ["충청북도", "충북"]),
    ("충청남도", ["충청남도", "충남"]),
    ("전북특별자치도", ["전북특별자치도", "전라북도", "전북"]),
    ("전라남도", ["전라남도", "전남"]),
    ("경상북도", ["경상북도", "경북"]),
    ("경상남도", ["경상남도", "경남"]),
    ("제주특별자치도", ["제주특별자치도", "제주도", "제주"]),
]

COLUMNS = ["사용가능여부", "중개인/임대인명", "창고명", "주소", "행정구역_도",
           "행정구역_시", "대지면적", "연면적", "건축면적", "준공연도",
           "공실현황", "시설특이사항", "기타", "정보확인일자", "사진"]
IMG_COLS = {"공실현황", "사진"}


# ---------------------------------------------------------------------------
# Geometry helpers
# ---------------------------------------------------------------------------
def _px(box: dict, w: int, h: int) -> tuple[int, int, int, int]:
    """Convert a fractional box into absolute pixel coordinates."""
    return (int(box["x0"] * w), int(box["y0"] * h),
            int(box["x1"] * w), int(box["y1"] * h))


def render_page(page: "fitz.Page", dpi: int) -> Image.Image:
    pix = page.get_pixmap(dpi=dpi)
    return Image.frombytes("RGB", (pix.width, pix.height), pix.samples)


# ---------------------------------------------------------------------------
# OCR helpers
# ---------------------------------------------------------------------------
def ocr(img: Image.Image, psm: int = 6, lang: str = OCR_LANG, scale: int = 1) -> str:
    if scale != 1:
        img = img.resize((img.width * scale, img.height * scale))
    return pytesseract.image_to_string(img, lang=lang, config=f"--psm {psm}").strip()


def detect_grid_rows(gray: np.ndarray, box_px: tuple[int, int, int, int],
                     dark_cut: int = 180, frac: float = 0.5) -> list[int]:
    """Return y-centres of horizontal grid lines inside the GI table box."""
    x0, y0, x1, y1 = box_px
    sub = gray[y0:y1, x0:x1]
    darkness = (sub < dark_cut).mean(axis=1)
    hits = [y0 + y for y, v in enumerate(darkness) if v > frac]
    groups: list[list[int]] = []
    for y in hits:
        if groups and y - groups[-1][-1] <= 4:
            groups[-1].append(y)
        else:
            groups.append([y])
    return [int(np.mean(g)) for g in groups]


# ---------------------------------------------------------------------------
# Value cleaners
# ---------------------------------------------------------------------------
def _strip_noise(text: str) -> str:
    text = re.sub(r"\s+", " ", text).strip()
    # OCR frequently appends the right cell border as stray glyphs.
    return text.strip(" :：|.,'\"`~^_-7ㆍ·")


_NUM = r"\d{1,3}(?:,\d{3})*(?:\.\d+)?"


def clean_area(text: str) -> str:
    """Normalise an area string like '29,70501 (8,986평)' -> '29,705㎡ (8,986평)'.

    The ㎡ figure and the (평) figure are read separately with a strict
    thousands/decimal pattern so OCR'd unit glyphs ('m', 'ni', '01', ...) that
    get fused onto the number are discarded.
    """
    left, _, right = text.partition("(")
    sqm_m = re.search(_NUM, left) or re.search(_NUM, text)
    pyeong_m = re.search(rf"({_NUM})\s*평", text) or re.search(_NUM, right)
    if sqm_m and pyeong_m:
        sqm = sqm_m.group(0).rstrip(".,")
        pyeong = (pyeong_m.group(1) if pyeong_m.re.groups else pyeong_m.group(0)).rstrip(".,")
        return f"{sqm}㎡ ({pyeong}평)"
    if sqm_m:
        return f"{sqm_m.group(0).rstrip('.,')}㎡"
    return _strip_noise(text)


def clean_year(text: str) -> str:
    m = re.search(r"((?:19|20)\d{2})\s*년\s*(\d{1,2})?\s*월?", text)
    if m:
        return f"{m.group(1)}년 {m.group(2)}월" if m.group(2) else f"{m.group(1)}년"
    m = re.search(r"(?:19|20)\d{2}", text)
    return f"{m.group(0)}년" if m else ""


def clean_title(text: str) -> str:
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"^\s*\d+\s*[.)]\s*", "", text)          # leading "1. "
    text = re.sub(r"\b[Nn][Ee][Ww]\b", "", text)            # red "NEW" badge
    text = text.strip(" .,")
    # Repair an unterminated "(상온" parenthesis from OCR.
    if text.count("(") > text.count(")"):
        text += ")"
    # Drop a parenthetical that OCR turned into pure garbage (no Hangul inside),
    # e.g. "물류센터 (%)" -> "물류센터".
    text = re.sub(r"\s*\((?![^)]*[가-힣])[^)]*\)", "", text)
    return text.strip()


def parse_region(address: str) -> tuple[str, str]:
    """Split an address into (도, 시/군). Handles spaced and unspaced forms."""
    addr = re.sub(r"\s+", "", address.strip())
    do = ""
    for canonical, forms in PROVINCE:
        matched = next((f for f in forms if addr.startswith(f)), None)
        if matched:
            do, addr = canonical, addr[len(matched):]
            break
    si = ""
    m = re.search(r"([가-힣]+?(?:시|군|구))", addr)
    if m:
        si = m.group(1)
    return do, si


# ---------------------------------------------------------------------------
# Image cropping
# ---------------------------------------------------------------------------
def content_bbox(img: Image.Image, box_px, white: int = 235, pad: int = 6):
    """Tighten a box to the non-white content it contains."""
    x0, y0, x1, y1 = box_px
    sub = np.array(img.crop((x0, y0, x1, y1)).convert("L"))
    mask = sub < white
    ys = np.where(mask.any(axis=1))[0]
    xs = np.where(mask.any(axis=0))[0]
    if len(xs) == 0 or len(ys) == 0:
        return box_px
    return (max(x0 + int(xs.min()) - pad, 0), max(y0 + int(ys.min()) - pad, 0),
            x0 + int(xs.max()) + pad, y0 + int(ys.max()) + pad)


# ---------------------------------------------------------------------------
# Per-page extraction
# ---------------------------------------------------------------------------
def extract_page(img: Image.Image, page_idx: int, img_dir: str) -> dict:
    w, h = img.size
    gray = np.array(img.convert("L"))
    gi_box = _px(GI_BOX, w, h)
    rows = detect_grid_rows(gray, gi_box)
    values: dict[str, str] = {}
    val_x0 = int(GI_VALUE_X * w)
    val_x1 = gi_box[2]
    for i in range(len(rows) - 1):
        if i >= len(GI_FIELDS):
            break
        yt, yb = rows[i] + 2, rows[i + 1] - 2
        if yb - yt < 8:
            continue
        cell = img.crop((val_x0, yt, val_x1, yb))
        values[GI_FIELDS[i]] = _strip_noise(ocr(cell, psm=7))

    address = values.get("소재지", "")
    do, si = parse_region(address)
    year = clean_year(values.get("준공년도", "")) or "미정"

    # --- Title (창고명) ----------------------------------------------------
    title_box = _px(TITLE_BOX, w, h)
    title = clean_title(ocr(img.crop(title_box), psm=7, lang="kor", scale=2))

    # --- 시설특이사항 (facility spec / 비고 column) ------------------------
    spec_raw = ocr(img.crop(_px(SPEC_BOX, w, h)), psm=6)
    spec_lines = [re.sub(r"^비고\s*", "", ln).strip()
                  for ln in spec_raw.splitlines() if ln.strip()]
    spec = "\n".join(spec_lines)

    # --- Images: building photo + Space-Availability table ----------------
    photo_box = content_bbox(img, _px(PHOTO_BOX, w, h))
    photo_path = os.path.join(img_dir, f"photo_p{page_idx + 1}.png")
    img.crop(photo_box).save(photo_path)

    space_path = os.path.join(img_dir, f"space_p{page_idx + 1}.png")
    img.crop(_px(SPACE_BOX, w, h)).save(space_path)

    return {
        "창고명": title,
        "주소": address,
        "행정구역_도": do,
        "행정구역_시": si,
        "대지면적": clean_area(values.get("대지면적", "")),
        "연면적": clean_area(values.get("연면적", "")),
        "준공연도": year,
        "시설특이사항": spec,
        "_photo": photo_path,
        "_space": space_path,
    }


# ---------------------------------------------------------------------------
# Text-layer extraction (for PDFs that already carry selectable text)
# ---------------------------------------------------------------------------
def is_property_page(text: str) -> bool:
    """A template property sheet has both tables, an address and a "N." title."""
    return ("General Information" in text and "Space Availability" in text
            and "소재지" in text and re.search(r"^\s*\d+\.\s*\S", text, re.M) is not None)


def _photo_rect(page: "fitz.Page"):
    """The building render is the largest embedded image in the top-left."""
    best, best_area = None, 0.0
    for im in page.get_images(full=True):
        for r in page.get_image_rects(im[0]):
            if r.x1 < 255 and r.y0 < 205:
                area = r.width * r.height
                if area > best_area and area > 8000:
                    best, best_area = r, area
    if best is None:  # fallback to a fixed left-column box
        best = fitz.Rect(32, 122, 233, 312)
    return best


def _space_rect(page: "fitz.Page"):
    """Box around the Space-Availability table: header down to the 합계 row."""
    sx = sy = total_bottom = None
    for w in page.get_text("words"):
        if w[4] == "Space":
            sx, sy = w[0], w[1]
        if w[4] == "합계":
            total_bottom = w[3]
    left = (sx - 6) if sx is not None else 388
    top = (sy - 8) if sy is not None else 310
    bottom = (total_bottom + 9) if total_bottom is not None else 495
    return fitz.Rect(left, top, page.rect.width - 6, bottom)


def _render_clip(page: "fitz.Page", rect, dpi: int, out_path: str) -> None:
    m = fitz.Matrix(dpi / 72, dpi / 72)
    page.get_pixmap(matrix=m, clip=rect).save(out_path)


def extract_text_page(page: "fitz.Page", page_idx: int,
                      img_dir: str, crop_dpi: int) -> dict:
    """Parse one property sheet from its text layer + render the two images."""
    text = page.get_text()
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    title = ""
    for ln in lines:
        if re.match(r"^\d+\.\s*\S", ln):
            title = clean_title(ln)
            break

    values: dict[str, str] = {}
    for i, ln in enumerate(lines):
        if ln in GI_FIELDS and i + 1 < len(lines):
            values.setdefault(ln, lines[i + 1])

    # 시설특이사항: the 비고 block, i.e. lines after "비고" up to the next label.
    spec = ""
    if "비고" in lines:
        bi = lines.index("비고")
        spec_lines = []
        for ln in lines[bi + 1:]:
            if ln in GI_FIELDS or ln in ("본부장", "부장", "대리", "Location"):
                break
            spec_lines.append(ln)
        spec = "\n".join(spec_lines)

    address = values.get("소재지", "")
    do, si = parse_region(address)
    year = clean_year(values.get("준공년도", "")) or "미정"

    photo_path = os.path.join(img_dir, f"photo_p{page_idx + 1}.png")
    _render_clip(page, _photo_rect(page), crop_dpi, photo_path)
    space_path = os.path.join(img_dir, f"space_p{page_idx + 1}.png")
    _render_clip(page, _space_rect(page), crop_dpi, space_path)

    return {
        "창고명": title,
        "주소": address,
        "행정구역_도": do,
        "행정구역_시": si,
        "대지면적": clean_area(values.get("대지면적", "")),
        "연면적": clean_area(values.get("연면적", "")),
        "준공연도": year,
        "시설특이사항": spec,
        "_photo": photo_path,
        "_space": space_path,
    }


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------
def px_to_col_width(px: int) -> float:
    return max(px / 7.0, 8)


def px_to_row_height(px: int) -> float:
    return px * 0.75  # px -> points (96dpi -> 72pt)


def write_excel(records: list[dict], out_path: str, broker: str,
                info_date: str, img_target_w: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouses"

    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    col_index = {name: i + 1 for i, name in enumerate(COLUMNS)}
    # Reasonable default widths for text columns.
    default_widths = {
        "사용가능여부": 12, "중개인/임대인명": 14, "창고명": 26, "주소": 26,
        "행정구역_도": 14, "행정구역_시": 12, "대지면적": 20, "연면적": 20,
        "건축면적": 14, "준공연도": 12, "시설특이사항": 38, "기타": 12,
        "정보확인일자": 14,
    }
    for name, width in default_widths.items():
        ws.column_dimensions[get_column_letter(col_index[name])].width = width

    max_row_px = 0
    for r, rec in enumerate(records, start=2):
        row_values = {
            "사용가능여부": "",
            "중개인/임대인명": broker,
            "창고명": rec["창고명"],
            "주소": rec["주소"],
            "행정구역_도": rec["행정구역_도"],
            "행정구역_시": rec["행정구역_시"],
            "대지면적": rec["대지면적"],
            "연면적": rec["연면적"],
            "건축면적": "",
            "준공연도": rec["준공연도"],
            "공실현황": "",
            "시설특이사항": rec["시설특이사항"],
            "기타": "",
            "정보확인일자": info_date,
            "사진": "",
        }
        for name, value in row_values.items():
            cell = ws.cell(row=r, column=col_index[name], value=value)
            cell.alignment = wrap_top
            cell.border = border

        # Embed the two images, scaled to a common target width.
        for name, path in (("공실현황", rec["_space"]), ("사진", rec["_photo"])):
            with Image.open(path) as im:
                ow, oh = im.size
            scale = img_target_w / ow
            sw, sh = int(ow * scale), int(oh * scale)
            xl = XLImage(path)
            xl.width, xl.height = sw, sh
            col_letter = get_column_letter(col_index[name])
            ws.column_dimensions[col_letter].width = px_to_col_width(sw)
            ws.add_image(xl, f"{col_letter}{r}")
            max_row_px = max(max_row_px, sh)

        ws.row_dimensions[r].height = px_to_row_height(max_row_px + 8)

    ws.freeze_panes = "A2"
    wb.save(out_path)


# ---------------------------------------------------------------------------
# Shared extraction driver (used by the CLI and the Streamlit app)
# ---------------------------------------------------------------------------
def ocr_available() -> bool:
    try:
        pytesseract.get_tesseract_version()
        return True
    except Exception:
        return False


def extract_records(doc: "fitz.Document", img_dir: str, dpi: int = 300,
                    img_dpi: int = 150, ocr_ready: bool | None = None,
                    progress=None) -> tuple[list[dict], int]:
    """Walk every page, returning (property records, skipped page count).

    `progress`, if given, is called as progress(page_index, total, record_or_None).
    """
    if ocr_ready is None:
        ocr_ready = ocr_available()
    os.makedirs(img_dir, exist_ok=True)

    records: list[dict] = []
    skipped = 0
    total = len(doc)
    for i, page in enumerate(doc):
        text = page.get_text()
        rec = None
        if is_property_page(text):
            rec = extract_text_page(page, i, img_dir, img_dpi)
        elif len(text.strip()) < 50 and ocr_ready:
            rec = extract_page(render_page(page, dpi), i, img_dir)
        else:
            skipped += 1
        if rec is not None:
            records.append(rec)
        if progress is not None:
            progress(i, total, rec)
    return records, skipped


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Extract MOVE logistics-center PDFs into an Excel sheet.")
    parser.add_argument("pdf", help="input PDF path")
    parser.add_argument("-o", "--output", help="output .xlsx path "
                        "(default: <pdf name>.xlsx)")
    parser.add_argument("--dpi", type=int, default=300,
                        help="render DPI for OCR on image-only PDFs (default: 300)")
    parser.add_argument("--img-dpi", type=int, default=150,
                        help="render DPI for cropped cell images (default: 150)")
    parser.add_argument("--broker", default="",
                        help="중개인/임대인명 value for every row (default: empty)")
    parser.add_argument("--date", dest="info_date",
                        help="정보확인일자 value (default: today as dd/mm/yyyy)")
    parser.add_argument("--img-width", type=int, default=360,
                        help="embedded image width in px (default: 360)")
    args = parser.parse_args(argv)

    if not os.path.isfile(args.pdf):
        print(f"error: file not found: {args.pdf}", file=sys.stderr)
        return 1

    out_path = args.output or os.path.splitext(args.pdf)[0] + ".xlsx"
    info_date = args.info_date or date.today().strftime("%d/%m/%Y")
    img_dir = os.path.splitext(out_path)[0] + "_images"
    os.makedirs(img_dir, exist_ok=True)

    doc = fitz.open(args.pdf)

    def _log(i, total, rec):
        if rec is not None:
            print(f"[page {i + 1}/{total}] {rec['창고명'] or '(no title)'} | "
                  f"{rec['행정구역_도']} {rec['행정구역_시']} | 준공 {rec['준공연도']}")

    records, skipped = extract_records(doc, img_dir, dpi=args.dpi,
                                       img_dpi=args.img_dpi, progress=_log)

    if not records:
        print("error: no property pages found in this PDF.", file=sys.stderr)
        return 3

    write_excel(records, out_path, args.broker, info_date, args.img_width)
    print(f"\nDone -> {out_path}  ({len(records)} rows, {skipped} non-pattern "
          f"pages skipped)")
    print(f"Cropped images -> {img_dir}/")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
